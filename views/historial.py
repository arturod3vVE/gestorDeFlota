import streamlit as st
import pandas as pd
import math
from io import BytesIO
from datetime import datetime, timedelta
from database import recuperar_historial_rango
from openpyxl.styles import Font, PatternFill

# --- DICCIONARIOS ---
DIAS_ES = {
    0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves", 
    4: "Viernes", 5: "Sábado", 6: "Domingo"
}
MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto", 
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

# --- FUNCIONES AUXILIARES (EXCEL) ---
def convertir_a_excel(datos_historial):
    filas = []
    for dia in datos_historial:
        fecha_str = dia['fecha']
        for asig in dia['reporte']:
            nombre_estacion = asig.get('nombre', 'Desconocido')
            horario = asig.get('horario', 'Sin horario')
            lista_unidades = asig.get('unidades', [])
            unidades_str = ", ".join([f"{u:02d}" for u in lista_unidades])
            cantidad = len(lista_unidades)
            
            filas.append({
                "Fecha": fecha_str,
                "Estación": nombre_estacion,
                "Horario": horario,
                "Total Unidades": cantidad,
                "Unidades Asignadas": unidades_str,
                "Última Edición": dia.get('actualizado', '')
            })
            
    if not filas: return None
    df = pd.DataFrame(filas)
    
    total_periodo = df["Total Unidades"].sum()
    fila_total = pd.DataFrame([{
        "Fecha": "TOTAL GLOBAL", "Estación": "", "Horario": "",
        "Total Unidades": total_periodo, "Unidades Asignadas": "", "Última Edición": ""
    }])
    df = pd.concat([df, fila_total], ignore_index=True)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reporte Mensual')
        worksheet = writer.sheets['Reporte Mensual']
        max_row = worksheet.max_row
        
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
            
        bold_font = Font(bold=True)
        fill_color = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=max_row, column=col)
            cell.font = bold_font
            cell.fill = fill_color
            
    output.seek(0)
    return output

# --- CALLBACKS DE PAGINACIÓN ---
def ir_a_pagina(n):
    st.session_state.pag_historial = n

def reset_pagina():
    st.session_state.pag_historial = 0

def render_vista(usuario_actual):
    st.title("📜 Reporte de Asignaciones")
    
    if 'pag_historial' not in st.session_state:
        st.session_state.pag_historial = 0
    
    # --- FILTRO ---
    c_filtro, c_metrics = st.columns([1, 2])
    with c_filtro:
        hoy = datetime.now()
        inicio_mes = hoy.replace(day=1)
        rango = st.date_input(
            "📅 Seleccionar Periodo",
            value=(inicio_mes, hoy),
            format="DD/MM/YYYY",
            on_change=reset_pagina 
        )
    
    st.divider()

    if isinstance(rango, tuple) and len(rango) == 2:
        inicio, fin = rango
        datos_completos = recuperar_historial_rango(usuario_actual, inicio, fin)
        
        if datos_completos:
            total_items = len(datos_completos)
            total_asig = sum([len(d["reporte"]) for d in datos_completos])
            
            with c_metrics:
                m1, m2, m3 = st.columns(3)
                m1.metric("Días Reportados", total_items)
                m2.metric("Total Asignaciones", total_asig)
                excel_file = convertir_a_excel(datos_completos)
                if excel_file:
                    nombre_archivo = f"Reporte_{inicio.strftime('%d%m')}_{fin.strftime('%d%m')}.xlsx"
                    m3.download_button("📥 Descargar Excel", excel_file, nombre_archivo, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
            
            st.divider()

            # --- LÓGICA DE PAGINACIÓN ---
            ITEMS_POR_PAGINA = 10
            total_paginas = math.ceil(total_items / ITEMS_POR_PAGINA)
            
            # Protección
            if st.session_state.pag_historial >= total_paginas:
                st.session_state.pag_historial = 0
            
            idx_inicio = st.session_state.pag_historial * ITEMS_POR_PAGINA
            idx_fin = idx_inicio + ITEMS_POR_PAGINA
            lote_actual = datos_completos[idx_inicio:idx_fin]
            
            st.caption(f"Mostrando {idx_inicio + 1} - {min(idx_fin, total_items)} de {total_items} registros")

            # --- RENDERIZADO DEL LOTE ---
            for i, item in enumerate(lote_actual):
                f_obj = datetime.strptime(item['fecha'], "%Y-%m-%d")
                nombre_dia = DIAS_ES[f_obj.weekday()]
                nombre_mes = MESES_ES[f_obj.month]
                fecha_bonita = f"{nombre_dia} {f_obj.day} de {nombre_mes}, {f_obj.year}"
                
                titulo_expander = f"📅 {fecha_bonita} ({len(item['reporte'])} asignaciones)"
                
                with st.expander(titulo_expander, expanded=False):
                    st.caption(f"📝 Creado: {item.get('creado', '--')} | 🔄 Última ed.: {item.get('actualizado', '--')}")
                    for asig in item['reporte']:
                        c_nom, c_uni = st.columns([2, 3])
                        c_nom.markdown(f"**⛽ {asig['nombre']}**")
                        if asig.get('horario'): c_nom.caption(f"🕒 {asig['horario']}")
                        
                        html_chips = "".join([f"<span style='background:#f0f2f6;padding:2px 6px;border-radius:4px;border:1px solid #ddd;font-weight:600;margin-right:4px;font-size:0.9em;display:inline-block;margin-bottom:2px;'>{u:02d}</span>" for u in asig['unidades']])
                        c_uni.markdown(html_chips, unsafe_allow_html=True)
                        st.markdown("<hr style='margin:5px 0;opacity:0.3'>", unsafe_allow_html=True)
                    
                    btn_col, _ = st.columns([1, 3])
                    unique_key = f"btn_edit_{item['fecha']}_{i}"
                    
                    if btn_col.button("✏️ Editar este reporte", key=unique_key):
                        st.session_state.reporte_diario = item['reporte']
                        st.session_state.fecha_reporte_activo = f_obj
                        st.session_state.vista_actual = "Asignacion"
                        st.rerun()

            # --- 🔘 PAGINADOR ESTILO GOOGLE 🔘 ---
            if total_paginas > 1:
                st.markdown("<br>", unsafe_allow_html=True)
                
                # 1. Configuración de Ventana
                MAX_BOTONES = 7
                pagina_actual = st.session_state.pag_historial
                
                inicio_ventana = max(0, pagina_actual - (MAX_BOTONES // 2))
                fin_ventana = min(total_paginas, inicio_ventana + MAX_BOTONES)
                
                if fin_ventana - inicio_ventana < MAX_BOTONES:
                    inicio_ventana = max(0, fin_ventana - MAX_BOTONES)

                # 2. PREPARAMOS LA LISTA DE BOTONES (Datos, no dibujo aún)
                botones_a_renderizar = []

                # Botón ANTERIOR
                if pagina_actual > 0:
                    botones_a_renderizar.append({
                        "label": "◀", "args": (pagina_actual - 1,), "type": "secondary", "key": "prev"
                    })

                # Botones NUMÉRICOS
                for p in range(inicio_ventana, fin_ventana):
                    tipo = "primary" if p == pagina_actual else "secondary"
                    botones_a_renderizar.append({
                        "label": str(p + 1), "args": (p,), "type": tipo, "key": f"pg_{p}"
                    })

                # Botón SIGUIENTE
                if pagina_actual < total_paginas - 1:
                    botones_a_renderizar.append({
                        "label": "▶", "args": (pagina_actual + 1,), "type": "secondary", "key": "next"
                    })

                # 3. EL TRUCO DE CENTRADO (COLUMNAS ESPACIADORAS)
                # Definimos ratios: 
                # [10] -> Espacio vacío izquierda (empuja al centro)
                # [1]... -> Cada botón ocupa poquito espacio
                # [10] -> Espacio vacío derecha
                
                cantidad_botones = len(botones_a_renderizar)
                ratio_columnas = [12] + [1] * cantidad_botones + [12]
                
                cols = st.columns(ratio_columnas)
                
                # Renderizamos los botones en las columnas centrales (saltando la primera)
                for i, btn_data in enumerate(botones_a_renderizar):
                    # Usamos cols[i + 1] porque cols[0] es el espaciador izquierdo
                    with cols[i + 1]:
                        st.button(
                            btn_data["label"],
                            key=f"btn_nav_{btn_data['key']}",
                            type=btn_data["type"],
                            on_click=ir_a_pagina,
                            args=btn_data["args"],
                            use_container_width=True # Esto hace que llenen su hueco pequeñito
                        )

        else:
            st.info("No se encontraron reportes en este rango de fechas.")
    else:
        st.warning("Selecciona una fecha de inicio y una de fin.")